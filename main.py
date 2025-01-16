import os
import httpx
import openpyxl
import json
import time
import random
import requests
import tkinter as tk
from tkinter import filedialog


def load_excel_file():
    """Загружает Excel файл и возвращает активный лист."""
    print("Откройте файл с данными...")
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename()
    file_name = file_path.split('/')[-1]
    file = file_name.rsplit('.', 1)[0]
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active
    root.destroy()
    return worksheet, file


def extract_data(worksheet, columns_to_extract):
    """Извлекает данные из указанных столбцов."""
    data = []
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number > 1:  # Пропуск первой строки
            extracted_row = [cell for idx, cell in enumerate(row) if idx + 1 in columns_to_extract]
            if any(extracted_row):  # Проверка на наличие данных
                data.append(extracted_row)
    print(data)
    return data


def get_user_input():
    """Получает ввод пользователя для поиска и режима работы."""
    site_name = input('Введите условие поиска (1 - без селлера, 2 - по селлеру): ')
    user_input = input("Введите режим '1 - Открытый', '2 - Скрытый': ")
    filtre_name = input('Введите условие поиска (1 - по популярности, 2 -по цене): ')

    filter = 'score'
    if filtre_name == '1':
        # код для поиска без селлера
        filter = 'score'
        print('Поиск по популярности:')
    elif filtre_name == '2':
        filter = 'price'
        # код для поиска по селлеру
        print('Поиск по цене:')
    else:
        print('Введите правильное число')
    headless = True if user_input.lower() == '2' else False
    mode = user_input  # Сохраняем режим для использования в других функциях
    return site_name, headless, mode, filter  # Возвращаем mode


def setup_proxies(mode):
    """Настраивает прокси в зависимости от режима."""
    return {
        '1': {},
        '2': {}
    }.get(mode, {})


def fetch_product_data(row, site_name, seller_numbers, mode, filter, count):
    keys_to_skip = []
    print(row)
    art = row[2]
    """Получает данные о продукте с сайта."""

    _HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36",
    }
    proxies = {
        '1': {},
        '2': {}
    }
    # Определение параметров запроса
    with httpx.Client(
            headers=_HEADERS,
            proxies=proxies.get(mode, None),
    ) as client:

        if seller_numbers:
            params = {
                'url': f'/search/?from_global=true&opened=seller&seller={seller_numbers}&text={row[3]}'
            }
        else:
            params = {
                'url': f'/search/?from_global=true&sorting={filter}&text={row[3]}'
            }

        try:
            response = client.get(
                'https://api.ozon.ru/entrypoint-api.bx/page/json/v2',
                params=params
            )

            data = response.json()

        except Exception as exc:
            print(f"Ошибка: {exc}")
    sku = ''
    widget_value = None
    skip_iteration = False
    widgets = data['widgetStates']
    sorted_widgets = sorted(widgets.items())
    sorted_widgets_dict = {widget_name: widget_value for widget_name, widget_value in sorted_widgets}
    for widget_name, widget_value in sorted_widgets_dict.items():
        widget_value = json.loads(widget_value)
        if 'fulltextResultsHeader-' in widget_name:
            try:
                header = widget_value['header']['text']
            except:
                header = ''
            print(header)
            if 'По вашему запросу товаров сейчас нет.' in header:

                print('-------------------------------------------------------------')
                skip_iteration = True
                break  # Выходим из цикла, если товаров нет.
        if 'searchResultsError-' in widget_name:
            print(widget_value['message'])
            if 'По вашим параметрам ничего не нашлось.' in widget_value['message']:
                skip_iteration = True
                print('-------------------------------------------------------------')
                break  # Выходим из цикла, если товаров нет.

        if 'filtersDesktop-' in widget_name:
            # Ищет фильтр 'Продавец' в данных.
            for section in widget_value['sections']:
                # Проверяем, есть ли 'Продавец' в текущем разделе
                if 'filters' in section:
                    for filter in section['filters']:
                        if filter['type'] == 'checkboxesFilter' and filter['key'] == 'seller':

                            items = filter['checkboxesFilter']['sections'][0]['items']
                            for item in items:
                                key = item['key']
                                if key in keys_to_skip:
                                    print('Найден товар селлера! Пропускаю товар.')
                                    continue  # пропускаем
                                # Обработать элемент далее, если это необходимо
                                # print(f'Обрабатываю товар селлера: {key}')
                else:
                    print("Нет 'filters' в in section:", section)

        if 'searchResultsV2-' in widget_name:
            items = widget_value['items']  # Получаем все элементы
            product = {}

            for item in items:  # Проходим по каждому элементу
                link = item['action']['link'].split('?')[0]
                link = 'https://ozon.ru' + link
                print(link)
                sku = link.split('-')[-1].replace('/', '')
                print(sku)
                product_name = [
                    c for c in item['mainState'] if ('id' in c.keys()) and (c['id'] == 'name')
                ] \
                    [0]['atom']['textAtom']['text'].replace('&#x2F;', '/').replace('&#34;', '"').replace('&#39;', "'")
                product_name_lower = product_name.lower()
                print(product_name_lower)
                part = str(row[2])  # преобразование part в строковый тип
                part_lower = part.lower() if part else ""
                print(part_lower)
                match = "Найдено" if part and part_lower in product_name_lower else "Не найдено"
                if match == "Найдено":
                    print(f'Совпадение part_номер в названии(Озон): {match}')
                    break


    return sku


def process_product_data(sku, mode, row, counter):
    """Обрабатывает данные о продукте и возвращает нужные параметры."""
    link = ''
    header = ''
    price_value = ''
    description1 = ''
    name = ''
    id = ''
    price = ''
    seller_name = ''
    seller_id = ''
    price_new = ''
    match = ''
    part_num = ''

    _HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 YaBrowser/24.1.0.0 Safari/537.36",
    }
    proxies = {
        '1': {},
        '2': {}
    }
    params = {

        'url': f'/product/{sku}'

    }
    with httpx.Client(
            headers=_HEADERS,
            proxies=proxies.get(mode, None),
    ) as client:
        try:
            response = client.get(
            'https://api.ozon.ru/entrypoint-api.bx/page/json/v2',
            params=params
            )

            data = response.json()

        except Exception as exc:
            print(f"Ошибка: {exc}")

    product = {}
    links_image = ''
    widgets = data['widgetStates']
    for widget_name, widget_value in widgets.items():
        widget_value = json.loads(widget_value)
        widgets = data['widgetStates']
        for widget_name, widget_value in widgets.items():
            # print(widget_name)
            widget_value = json.loads(widget_value)
            if 'webGallery-' in widget_name:
                try:
                    image_urls = []
                    for item in widget_value['images']:
                        image_urls.append(item['src'])
                    links_image = ' '.join(image_urls)
                    # print(image_urls)
                except KeyError:
                    links_image = ''
                    print("No 'images' key found in widget_value")
            if 'webStickyProducts' in widget_name:
                # print(widget_name)
                try:
                    product['name'] = widget_value['name']
                    name = widget_value['name']
                except:
                    name = ''
                # print(name)
                try:
                    product['sku'] = widget_value['sku']
                    id = widget_value['sku']
                except:
                    id = ''
                # print(id)
                try:
                    if widget_value.get('seller', {}).get('link'):
                        product['seller'] = widget_value['seller']['link'].split('/')[2]
                        seller_id = product['seller']
                    else:
                        seller_id = ''
                except:
                    seller_id = ''
                # print(product['seller'])
                try:
                    seller_name = widget_value['seller']['name']
                except:
                    seller_name = ''
                # print(seller_name)
            try:
                if ("webPrice" in widget_name and
                        "isAvailable" in widget_value.keys()):
                    if widget_value['isAvailable'] == True:
                        product["price"] = widget_value["price"]
                        price = widget_value["price"].replace('₽', '').replace('\u2009', '')
                        price = int(price)
            except:
                price = ''


    yandex = ''
    with httpx.Client(
            headers=_HEADERS,
            proxies=proxies.get(mode, None),
    ) as client:

        try:

            params = {

                'url': f'/product/{sku}?layout_container=pdpPage2column&layout_page_index=2'

            }
            response = client.get(
                'https://api.ozon.ru/entrypoint-api.bx/page/json/v2',
                params=params
            )
            print(response)

            data = response.json()

        except Exception as exc:
            print(f"Ошибка: {exc}")



    product = {}
    widgets = data['widgetStates']
    for widget_name, widget_value in widgets.items():
        widget_value = json.loads(widget_value)
        # print(widget_name)
        if 'webCharacteristics-' in widget_name:
            products = widget_value['characteristics']
            product['characteristics'] = widget_value['characteristics']
            for characteristic in product['characteristics']:
                # print(characteristic)
                try:
                    for value in characteristic['short']:
                        # print(value)
                        if value['key'] == 'Model':
                            part_num = value['values'][0]['text']
                            # print('Part_Num:', part_num)
                            product['part_num'] = part_num
                            print('Part_Num:', product['part_num'])
                            # break
                        else:
                            product['part_num'] = ''
                except:
                    continue
    link = 'https://ozon.ru' + sku
    print(
        f"{counter}_Товар: NAME: {name}, PART_NUMBER: {part_num}, ID: {id}, SELLER_NAME: {seller_name}, SELLER: {seller_id} PRICE: {price} LINK: {link} 'Картинки:', {links_image}")
    if seller_id in []:
        print('Найден товар селлера! Пропускаю товар.')


    print('-------------------------------------------------------------')

    data_row = [
        row[0],
        row[1],
        row[2],
        row[3],
        price_new,
        price,
        name,
        match,
        yandex,
        sku,
        part_num,
        seller_id,
        seller_name,
        link,
        links_image
    ]


    return data_row


def initialize_workbook(file_path):
    """Инициализирует рабочую книгу, загружая существующий файл или создавая новый."""
    if os.path.exists(file_path):
        # Загружаем существующий файл
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    else:
        # Создаем новый файл
        wb = openpyxl.Workbook()
        ws = wb.active
        # Записываем заголовки только если создаем новый файл
        header_row = [
            "Код 1c", "Брэнд", "Part_номер", "Название", "Цена",
            "Цена(Озон)", "Название(Озон)", "Совпадение", "Поиск(Яндекс)",
            "id(Озон)", "Part_номер(Озон)", "Селлер_id(Озон)",
            "Селлер(Озон)", "Ссылка(Озон)", "Картинки(Озон)"
        ]
        ws.append(header_row)

    return wb, ws


def save_results_to_excel(data, file):
    """Сохраняет результаты в новый Excel файл или добавляет в существующий."""
    file_path = f"{file}_Result.xlsx"

    # Инициализация рабочей книги и активного листа
    wb, ws = initialize_workbook(file_path)

    # Добавляем каждую строку данных
    for row in data:
        ws.append(row)

    # Сохраняем книгу
    wb.save(file_path)


def main():
    """Основная функция для запуска программы."""
    worksheet, file = load_excel_file()
    data = extract_data(worksheet, [1, 2, 3, 4, 5])
    site_name, headless, mode, filter = get_user_input()

    # Дополнительные настройки
    seller_numbers = []  # Здесь можно указать номера селлеров, если необходимо
    # filter = 'score'  # Или 'price', в зависимости от ввода пользователя
    result_data = []

    count = 1
    counter = 1
    for row in data:
        product_data = fetch_product_data(row, site_name, seller_numbers, headless, filter, count)
        count += 1
        if product_data:
            print(counter)
            processed_data = process_product_data(product_data, mode, row, counter)
            print(processed_data)
            counter += 1
            ## Сохраняем
            save_results_to_excel([processed_data], file)
    # Сохранение результатов в Excel
    save_results_to_excel(result_data, file)
    print(f"Обработка завершена. Результаты сохранены в {file}_Result.xlsx'.")

if __name__ == "__main__":
    main()
